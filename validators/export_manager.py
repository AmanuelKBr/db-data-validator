import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import re
import csv
from io import StringIO, BytesIO
from typing import Union
from utils.models import TableValidationReport, BatchValidationReport
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HDR_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

_SEV_BG = {"CRITICAL": "FFE6E6", "WARNING": "FFF2CC", "INFO": "E7F0FF"}
_SEV_FG = {"CRITICAL": "C00000", "WARNING": "9C6500", "INFO": "0070C0"}


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN


def _style_data_row(ws, row_num, severity_value):
    bg = _SEV_BG.get(severity_value, "FFFFFF")
    fg = _SEV_FG.get(severity_value, "000000")
    for cell in ws[row_num]:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(color=fg, bold=(severity_value == "CRITICAL"))
        cell.border = _THIN
        cell.alignment = _CELL_ALIGN


def _friendly_expected(rule_violated: str) -> str:
    """Translate SQL validation logic into a human-readable expected condition.

    Checks are ordered most-specific → least-specific so that compound rules
    like '[COL] IS NULL OR [COL] IN (...)' are handled correctly instead of
    being short-circuited by the IS NULL branch.
    """
    if not rule_violated:
        return "—"

    rv       = rule_violated.strip()
    rv_upper = rv.upper()

    # ── 1. Allowed-value list: IN (...) — check BEFORE IS NULL ────────────
    m = re.search(r'\bIN\s*\(([^)]+)\)', rv, re.IGNORECASE)
    if m:
        raw    = m.group(1).split(",")
        values = [v.strip().strip("'\"") for v in raw if v.strip().strip("'\"")]
        preview = ", ".join(values[:8])
        suffix  = "…" if len(values) > 8 else ""
        return f"One of: {preview}{suffix}"

    # ── 2. Casing / pattern checks ────────────────────────────────────────
    if "PATINDEX" in rv_upper or "COLLATE" in rv_upper:
        return "Proper name casing (no irregular mixed case)"

    # ── 3. Date checks ────────────────────────────────────────────────────
    if "GETDATE()" in rv_upper or "GETUTCDATE()" in rv_upper:
        return "Date ≤ today (not in future)"
    if "DATEADD" in rv_upper or "DATEDIFF" in rv_upper:
        return "Date within valid range"
    if "2000" in rv:
        return "Date ≥ 2000-01-01"
    if "1900" in rv:
        return "Date ≥ 1900-01-01"

    # ── 4. Duplicate check ────────────────────────────────────────────────
    if "GROUP BY" in rv_upper and "HAVING" in rv_upper:
        return "Unique (no duplicates)"

    # ── 5. Numeric range ──────────────────────────────────────────────────
    if "BETWEEN" in rv_upper:
        m2 = re.search(r'BETWEEN\s+([\d\.]+)\s+AND\s+([\d\.]+)', rv, re.IGNORECASE)
        return f"Between {m2.group(1)} and {m2.group(2)}" if m2 else "Value within range"
    if ">= 0 AND" in rv_upper and "<=" in rv_upper:
        m3 = re.search(r'<=\s*(\d[\d,\.]*)', rv)
        return f"0 ≤ value ≤ {m3.group(1)}" if m3 else "Value within allowed range"
    if ">= 0" in rv_upper:
        return "Value ≥ 0 (non-negative)"
    if "> 0" in rv_upper:
        return "Value > 0 (positive)"
    if "< 0" in rv_upper:
        return "Value < 0"
    if "<> 0" in rv_upper or "!= 0" in rv_upper:
        return "Value ≠ 0"

    # ── 6. String checks ──────────────────────────────────────────────────
    if "LTRIM" in rv_upper or "RTRIM" in rv_upper:
        return "Non-blank (no whitespace-only values)"
    if "= ''" in rv_upper:
        return "Non-empty string"
    if "LEN(" in rv_upper:
        m4 = re.search(r'LEN\([^\)]+\)\s*<=\s*(\d+)', rv, re.IGNORECASE)
        if m4:
            return f"Length ≤ {m4.group(1)} characters"
        m5 = re.search(r'LEN\([^\)]+\)\s*>=\s*(\d+)', rv, re.IGNORECASE)
        if m5:
            return f"Length ≥ {m5.group(1)} characters"
        return "Valid string length"
    if "LIKE" in rv_upper:
        m6 = re.search(r"LIKE\s+'([^']+)'", rv, re.IGNORECASE)
        return f"Matches pattern: {m6.group(1)}" if m6 else "Matches required format"

    # ── 7. Format conversion ──────────────────────────────────────────────
    if "TRY_CONVERT" in rv_upper:
        return "Valid numeric/date format"

    # ── 8. NULL checks — last, because many rules contain IS NULL as part
    #       of nullable-column logic (e.g. IS NULL OR <condition>) ─────────
    if "IS NOT NULL" in rv_upper:
        return "Value must be present (NOT NULL)"
    if "IS NULL" in rv_upper:
        return "NOT NULL (value required)"

    # ── 9. Fallback ───────────────────────────────────────────────────────
    return f"Satisfies: {rv[:80]}"


class ExportManager:
    """Handles export of validation reports to CSV and Excel formats."""

    # ── CSV exports ──────────────────────────────────────────────────────────

    @staticmethod
    def export_table_report_to_csv(report: TableValidationReport) -> str:
        out = StringIO()
        w = csv.writer(out)

        w.writerow([f"DATA QUALITY VALIDATION REPORT — {report.table_name}"])
        w.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        w.writerow([f"Total rows checked: {report.total_rows_checked:,}"])
        w.writerow([])
        w.writerow(["=== SUMMARY (one row per rule) ==="])
        w.writerow(["Table", "Rule", "Column", "Severity", "Total Failures", "Failure %"])
        for r in report.results:
            w.writerow([
                report.table_name, r.rule_name, r.column_name,
                r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%",
            ])

        w.writerow([])
        total_failure_rows = sum(len(r.failures) for r in report.results)
        w.writerow([f"=== FAILURE DETAILS — {total_failure_rows:,} individual rows (up to 1,000 per rule) ==="])
        w.writerow(["Table", "Rule", "Column", "Severity", "Row ID", "Actual Value", "Expected Condition"])
        for r in report.results:
            expected = _friendly_expected(r.failures[0].rule_violated if r.failures else "")
            for f in r.failures:
                w.writerow([
                    report.table_name, r.rule_name, f.column_name, r.severity.value,
                    f.row_id,
                    str(f.actual_value) if f.actual_value is not None else "NULL",
                    expected,
                ])
        return out.getvalue()

    @staticmethod
    def export_batch_report_to_csv(batch_report: BatchValidationReport) -> str:
        out = StringIO()
        w = csv.writer(out)

        w.writerow(["BATCH VALIDATION REPORT"])
        w.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        w.writerow(["Tables validated", batch_report.total_tables])
        w.writerow(["Total issues", batch_report.total_issues])
        w.writerow(["Critical", batch_report.get_critical_count()])
        w.writerow(["Warning", batch_report.get_warning_count()])
        w.writerow(["Info", batch_report.get_info_count()])
        w.writerow([])
        w.writerow(["=== SUMMARY (one row per rule, all tables) ==="])
        w.writerow(["Table", "Rule", "Column", "Severity", "Total Failures", "Failure %"])
        for tr in batch_report.table_reports:
            for r in tr.results:
                w.writerow([
                    tr.table_name, r.rule_name, r.column_name,
                    r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%",
                ])

        w.writerow([])
        total_rows = sum(len(r.failures) for tr in batch_report.table_reports for r in tr.results)
        w.writerow([f"=== FAILURE DETAILS — {total_rows:,} individual rows ==="])
        w.writerow(["Table", "Rule", "Column", "Severity", "Row ID", "Actual Value", "Expected Condition"])
        for tr in batch_report.table_reports:
            for r in tr.results:
                expected = _friendly_expected(r.failures[0].rule_violated if r.failures else "")
                for f in r.failures:
                    w.writerow([
                        tr.table_name, r.rule_name, f.column_name, r.severity.value,
                        f.row_id,
                        str(f.actual_value) if f.actual_value is not None else "NULL",
                        expected,
                    ])
        return out.getvalue()

    # ── Excel exports ─────────────────────────────────────────────────────────

    @staticmethod
    def export_table_report_to_excel(report: TableValidationReport) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Summary")
        ws.append(["Data Quality Validation Report"])
        ws.append([f"Table: {report.table_name}"])
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Total rows checked: {report.total_rows_checked:,}"])
        ws.append([])
        ws.append(["Rule", "Column", "Severity", "Total Failures", "Failure %"])
        _style_header(ws)
        for r in report.results:
            ws.append([r.rule_name, r.column_name, r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%"])
            _style_data_row(ws, ws.max_row, r.severity.value)
        for col, w in zip("ABCDE", [30, 20, 12, 18, 12]):
            ws.column_dimensions[col].width = w

        wd = wb.create_sheet("All Failures")
        wd.append(["Rule", "Column", "Severity", "Row ID", "Actual Value", "Expected Condition"])
        _style_header(wd)
        for r in report.results:
            expected = _friendly_expected(r.failures[0].rule_violated if r.failures else "")
            for f in r.failures:
                wd.append([
                    r.rule_name, f.column_name, r.severity.value,
                    f.row_id,
                    str(f.actual_value) if f.actual_value is not None else "NULL",
                    expected,
                ])
                _style_data_row(wd, wd.max_row, r.severity.value)
        wd.append([])
        note_row = wd.max_row + 1
        wd.cell(note_row, 1, "Note: up to 1,000 rows captured per rule. Total failures may be higher — see Summary sheet.")
        for col, w in zip("ABCDEF", [30, 20, 12, 10, 35, 40]):
            wd.column_dimensions[col].width = w

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def export_batch_report_to_excel(batch_report: BatchValidationReport) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["Batch Validation Report"])
        ws_sum.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_sum.append([])
        ws_sum.append(["Tables Validated", batch_report.total_tables])
        ws_sum.append(["Total Issues", batch_report.total_issues])
        ws_sum.append(["Critical", batch_report.get_critical_count()])
        ws_sum.append(["Warning", batch_report.get_warning_count()])
        ws_sum.append(["Info", batch_report.get_info_count()])
        ws_sum.append([])
        ws_sum.append(["Table", "Rule", "Column", "Severity", "Total Failures", "Failure %"])
        _style_header(ws_sum)
        for tr in batch_report.table_reports:
            for r in tr.results:
                ws_sum.append([tr.table_name, r.rule_name, r.column_name, r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%"])
                _style_data_row(ws_sum, ws_sum.max_row, r.severity.value)
        for col, w in zip("ABCDEF", [25, 28, 20, 12, 18, 12]):
            ws_sum.column_dimensions[col].width = w

        wd = wb.create_sheet("All Failures")
        wd.append(["Table", "Rule", "Column", "Severity", "Row ID", "Actual Value", "Expected Condition"])
        _style_header(wd)
        for tr in batch_report.table_reports:
            for r in tr.results:
                expected = _friendly_expected(r.failures[0].rule_violated if r.failures else "")
                for f in r.failures:
                    wd.append([
                        tr.table_name, r.rule_name, f.column_name, r.severity.value,
                        f.row_id,
                        str(f.actual_value) if f.actual_value is not None else "NULL",
                        expected,
                    ])
                    _style_data_row(wd, wd.max_row, r.severity.value)
        for col, w in zip("ABCDEFG", [25, 28, 20, 12, 10, 35, 40]):
            wd.column_dimensions[col].width = w

        for tr in batch_report.table_reports:
            sname = tr.table_name.replace(".", "_")[:31]
            ws = wb.create_sheet(sname)
            ws.append([f"Table: {tr.table_name}  |  {tr.total_rows_checked:,} rows  |  {len(tr.results)} issues"])
            ws.append([])
            ws.append(["Rule", "Column", "Severity", "Total Failures", "Failure %", "Row ID", "Actual Value", "Expected Condition"])
            _style_header(ws)
            for r in tr.results:
                expected = _friendly_expected(r.failures[0].rule_violated if r.failures else "")
                if r.failures:
                    for f in r.failures:
                        ws.append([r.rule_name, f.column_name, r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%", f.row_id, str(f.actual_value) if f.actual_value is not None else "NULL", expected])
                        _style_data_row(ws, ws.max_row, r.severity.value)
                else:
                    ws.append([r.rule_name, r.column_name, r.severity.value, r.total_failures, f"{r.failure_percentage:.2f}%", "—", "—", expected])
                    _style_data_row(ws, ws.max_row, r.severity.value)
            for col, w in zip("ABCDEFGH", [28, 20, 12, 18, 12, 10, 35, 40]):
                ws.column_dimensions[col].width = w

        out = BytesIO()
        wb.save(out)
        return out.getvalue()